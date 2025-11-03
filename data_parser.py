import csv
import ast
import openpyxl
from openpyxl import load_workbook

    
def write_road_info(path,data,clear_file = False):#clear_file to be usable if schema changed
    if not data:
        return  # nothing to write
    
    mode = "w" if clear_file else "a"
    fieldnames = list(data[0].keys())  # dynamically get headers from first row

    with open(path, mode=mode, newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)

        # Only write header if file is empty
        if f.tell() == 0:
            writer.writeheader()

        for row in data:
            processed_row = {}
            for key, value in row.items():
                if isinstance(value, list):
                    processed_row[key] = ";".join(map(str, value))
                else:
                    processed_row[key] = value
            writer.writerow(processed_row)





def parse_value(val):
    """Try to parse into int, float, list, or keep as string."""
    if ";" in val:  # handle semicolon-separated lists
        parts = val.split(";")
        try:
            return [int(p) for p in parts]
        except ValueError:
            try:
                return [float(p) for p in parts]
            except ValueError:
                return parts
    else:
        try:
            return int(val)
        except ValueError:
            try:
                return float(val)
            except ValueError:
                try:
                    return ast.literal_eval(val)  # catch Python literals like lists
                except Exception:
                    return val

def read_road_info(path, match_value, match_column="episode"):
    """
    Reads CSV and returns parsed dicts only for rows where `match_column` == match_value.

    Args:
        path (str): Path to CSV file.
        match_value (any): Value to filter for.
        match_column (str): Column name to match.
    """
    data = []
    with open(path, mode="r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if str(row[match_column]) == str(match_value):
                parsed_row = {k: parse_value(v) for k, v in row.items()}
                data.append(parsed_row)
    return data



def ReadRow(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]  # first row = headers
    
    for row in ws.iter_rows(min_row=2, values_only=False):  # skip header
        row_dict = {headers[i]: row[i].value for i in range(len(headers))}
        
        if row_dict.get("Waiting Time (s)") is None:  # empty row found
            # return only non-empty values
            return {k: v for k, v in row_dict.items() if v is not None}
    
    return None  # if no row found


def WriteRow(data_dict, file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]  # first row = headers
    
    for row in ws.iter_rows(min_row=2):  # go through rows
        row_values = {headers[i]: row[i].value for i in range(len(headers))}
        
        if row_values.get("Waiting Time (s)") is None:  # first empty row
            for key, value in data_dict.items():
                if key in headers:
                    col_idx = headers.index(key) + 1
                    ws.cell(row=row[0].row, column=col_idx, value=value)
            break

    wb.save(file_path)

def count_full_rows(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # get headers
    headers = [cell.value for cell in ws[1]]
    
    # find column index for "Waiting Time (s)"
    try:
        col_idx = headers.index("Waiting Time (s)") + 1  # +1 for 1-based indexing
    except ValueError:
        raise ValueError("Column 'Waiting Time (s)' not found in sheet")

    count = 0
    for row in ws.iter_rows(min_row=2):  # skip header row
        if row[col_idx - 1].value is not None:
            count += 1

    return count

def clean_dict_values(dict_):
    if "Area" in dict_:
          dict_["Area"] = 'Mosheer' if ('Mosheer' in dict_["Area"]) else 'Stefano' 
    if "Reward" in dict_:
          dict_["Reward"] = 'proposed_reward' if ('Proposed' in dict_["Reward"]) else 'literature'
    if "Traffic Scale" in dict_:
          dict_["Traffic Scale"] = .14 if ('Normal' in dict_["Traffic Scale"]) else .38
    return dict_



